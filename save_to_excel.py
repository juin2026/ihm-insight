import io
import pandas as pd
from datetime import datetime
from pathlib import Path

try:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

DISPLAY_COLS = [
    "날짜", "미디어타입", "캡션", "링크",
    "도달", "좋아요", "댓글", "저장", "공유", "조회수",
    "참여율(%)", "공유율(%)", "반복시청률(%)",
]


def _style_sheet(ws):
    if not _HAS_OPENPYXL:
        return
    header_fill = PatternFill(start_color="1F2D3D", end_color="1F2D3D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[letter].width = min(max(max_len + 3, 8), 50)


def _build_summary(df: pd.DataFrame) -> pd.DataFrame:
    rv = df[df["미디어타입_원본"].isin(["REELS", "VIDEO"])]
    fd = df[df["미디어타입_원본"].isin(["IMAGE", "CAROUSEL_ALBUM"])]

    def _stats(sub: pd.DataFrame) -> list:
        return [
            len(sub),
            sub["도달"].sum(),
            sub["좋아요"].sum(),
            sub["댓글"].sum(),
            sub["저장"].sum(),
            sub["공유"].sum(),
            sub["조회수"].sum(),
            round(sub["참여율(%)"].mean(), 2) if len(sub) else 0,
            round(sub["공유율(%)"].mean(), 2) if len(sub) else 0,
            round(sub["반복시청률(%)"].mean(), 2) if len(sub) else 0,
        ]

    labels = [
        "총 게시물", "총 도달", "총 좋아요", "총 댓글", "총 저장", "총 공유",
        "총 조회수",
        "평균 참여율(%)", "평균 공유율(%)", "평균 반복시청률(%)",
    ]

    return pd.DataFrame({
        "지표": labels,
        "전체": _stats(df),
        "릴스·영상": _stats(rv),
        "피드 이미지·캐러셀": _stats(fd),
    })


def _write_excel(df: pd.DataFrame, target) -> None:
    """Write DataFrame to target (file path str or BytesIO)."""
    disp = [c for c in DISPLAY_COLS if c in df.columns]
    rv_df = df[df["미디어타입_원본"].isin(["REELS", "VIDEO"])][disp]
    fd_df = df[df["미디어타입_원본"].isin(["IMAGE", "CAROUSEL_ALBUM"])][disp]
    summary_df = _build_summary(df)

    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        rv_df.to_excel(writer, sheet_name="릴스·영상", index=False)
        fd_df.to_excel(writer, sheet_name="피드 이미지·캐러셀", index=False)
        summary_df.to_excel(writer, sheet_name="전체 요약", index=False)

        for sheet_name in writer.sheets:
            _style_sheet(writer.sheets[sheet_name])


def save_to_excel(df: pd.DataFrame, output_dir: str = ".") -> str:
    """Save DataFrame to a dated Excel file. Returns the saved file path."""
    today = datetime.now().strftime("%Y%m%d")
    filename = f"{today}_아이헤이트먼데이인사이트.xlsx"
    filepath = Path(output_dir) / filename
    _write_excel(df, str(filepath))
    return str(filepath)


def get_excel_bytes(df: pd.DataFrame) -> bytes:
    """Return Excel file as bytes for Streamlit download_button."""
    buf = io.BytesIO()
    _write_excel(df, buf)
    return buf.getvalue()


if __name__ == "__main__":
    from fetch_insights import fetch_insights
    df = fetch_insights()
    path = save_to_excel(df)
    print(f"저장 완료: {path}")
